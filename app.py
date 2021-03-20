#----------------------------------------------------------------------------#
# Imports
#----------------------------------------------------------------------------#

import uuid

import flask
from flask import Flask, render_template, request
from flask_sqlalchemy import SQLAlchemy
import logging
from logging import Formatter, FileHandler
from forms import *
from openpyxl import load_workbook
import os
import re
import json


#----------------------------------------------------------------------------#
# App Config.
#----------------------------------------------------------------------------#

app = Flask(__name__)
app.config.from_object('config')
db = SQLAlchemy(app)

# Automatically tear down SQLAlchemy.
'''
@app.teardown_request
def shutdown_session(exception=None):
    db_session.remove()
'''

# Login required decorator.
'''
def login_required(test):
    @wraps(test)
    def wrap(*args, **kwargs):
        if 'logged_in' in session:
            return test(*args, **kwargs)
        else:
            flash('You need to login first.')
            return redirect(url_for('login'))
    return wrap
'''
#----------------------------------------------------------------------------#
# Controllers.
#----------------------------------------------------------------------------#


@app.route('/')
def home():
    if 'uid' not in flask.session:
        flask.session['id'] = str(uuid.uuid4())
    return render_template('pages/placeholder.home.html')


@app.route('/login')
def login():
    form = LoginForm(request.form)
    return render_template('forms/login.html', form=form)


@app.route('/register')
def register():
    form = RegisterForm(request.form)
    return render_template('forms/register.html', form=form)


@app.route('/forgot')
def forgot():
    form = ForgotForm(request.form)
    return render_template('forms/forgot.html', form=form)


@app.route('/shipments', methods=['GET'])
def shipments():
    filename = flask.session['id']
    '''
    if not os.path.contains():
        flash("Session not found")
    '''
    return render_template('pages/upload_shipment.html')

@app.route('/shipments/scan', methods=['GET'])
def scan():
    filename = flask.session['id']
    '''
    if not os.path.contains():
        flash("Session not found")
    '''
    return render_template('pages/scan_item.html')

@app.route('/shipments/pack', methods=['GET'])
def shipments_pack():
    # get count of units
    # return to template
    data = {
        "shipment_id": flask.session["shipment_id"],
        "unit_count": flask.session["unit_count"]
    }
    return render_template('pages/pack_shipment.html', data=data)

@app.route('/shipments/complete', methods=['POST'])
def complete_shipments():
    boxes = request.get_json()
    filename = os.path.join(app.config['UPLOAD_FOLDER'], flask.session['id'])+ ".xlsx"
    wb = load_workbook(filename=filename)
    ws = wb.get_sheet_by_name("Pack List")

    for box in boxes:
        box_id = box["box_number"]
        weight = box["weight"]
        length = box["length"]
        width = box["width"]
        height = box["height"]

        for item in box["items"]:
            upc = item["UPC"]
            quantity = item["quantity"]

            for cell in ws['E']:
                if(cell.value is not None) and upc in cell.value:
                    cell_quantity = ws.cell(row=cell.row,column=12+box_id).value or 0
                    cell_quantity += quantity
                    ws.cell(row=cell.row,column=12+box_id).value = cell_quantity
                    break

            for cell in ws['C']:
                if(cell.value is not None):
                    if "Weight" in cell.value:
                        ws.cell(row=cell.row,column=12+box_id).value = weight
                    if "length" in cell.value:
                        ws.cell(row=cell.row,column=12+box_id).value = length
                    if "width" in cell.value:
                        ws.cell(row=cell.row,column=12+box_id).value = width
                    if "height" in cell.value:
                        ws.cell(row=cell.row,column=12+box_id).value = height

    wb.save(filename)
    return json.dumps({'success':True}), 200, {'ContentType':'application/json'}

@app.route('/shipments', methods=['POST'])
def upload_shipments():
    fba_file = request.files['file']
    extension = fba_file.filename[fba_file.filename.rfind("."):]
    # save to disk with session id
    filename = os.path.join(app.config['UPLOAD_FOLDER'], flask.session['id']) + ".xlsx"
    fba_file.save(filename)
    # count number of units in excel
    wb = load_workbook(filename=filename)
    ws = wb.get_sheet_by_name("Pack List")
    pattern = "Total Units: (?P<units>[0-9]+)$"

    for cell in ws['A']:
        if(cell.value is not None) and "Total Units" in cell.value:
                for match in re.finditer(pattern, cell.value):
                    units = int(match["units"])
                break
    flask.session["unit_count"] = units # set unit count
    flask.session["shipment_id"] = ws['B'][0].value

    return flask.redirect('/shipments/pack')

@app.route('/shipments/download', methods=['GET'])
def download_shipment():
    disk_filename = "fba/" + flask.session["id"] + ".xlsx"
    serv_filename = flask.session["shipment_id"] + ".xlsx"
    return flask.send_file(disk_filename, attachment_filename=serv_filename, as_attachment=True)


# Error handlers.
@app.errorhandler(500)
def internal_error(error):
    #db_session.rollback()
    return render_template('errors/500.html'), 500


@app.errorhandler(404)
def not_found_error(error):
    return render_template('errors/404.html'), 404

if not app.debug:
    file_handler = FileHandler('error.log')
    file_handler.setFormatter(
        Formatter('%(asctime)s %(levelname)s: %(message)s [in %(pathname)s:%(lineno)d]')
    )
    app.logger.setLevel(logging.INFO)
    file_handler.setLevel(logging.INFO)
    app.logger.addHandler(file_handler)
    app.logger.info('errors')

#----------------------------------------------------------------------------#
# Launch.
#----------------------------------------------------------------------------#

# Default port:
if __name__ == '__main__':
    app.run()

# Or specify port manually:
'''
if __name__ == '__main__':
    port = int(os.environ.get('PORT', 5000))
    app.run(host='0.0.0.0', port=port)
'''
